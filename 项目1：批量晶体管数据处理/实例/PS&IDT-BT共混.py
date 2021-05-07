from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import os
import math


def load_files(n):
    '''

    :param path:
    :return: files_xlsx
    '''
    # 读取路径默认，需要在调用时手动更改
    # 重复读取excel文件，并依次交给 load_excel 函数执行
    # 数据文件存储路径
    path = '/Users/mac/Documents/MyData/PS IDT-BT共混/' + str(n)
    # 读取 path 下的所有文件的文件名
    files_name = os.listdir(path)
    files_name.sort()

    return files_name


def load_excel(n, file_name, i=82, colx=7, coly=5):
    '''

    :param path:
    :param i:
    :param colx:
    :param coly:
    :return: CH1_Current, CH2_Voltage
    '''
    # 数据文件存储路径
    path = '/Users/mac/Documents/MyData/PS IDT-BT共混/' + str(n) + '/' + file_name
    # 导入excel内所需要的数据
    # path 数据存放路径；i 导入几行(从第二行开始计数);
    # colx 导入的作为x轴的数据的列号;coly 导入的作为y轴的数据的列号
    # 部分数据存在少于81组数据的情况，需要修改 i 这部分的数值。不需要修改，
    # 这部分少数据大概率是因为软件导出到 excel 表格的时候出错，导致部分数据缺失了

    a = load_workbook(path)
    A = a.active

    CH1_Current = []
    CH2_Voltage = []

    for j in range(1, i + 1):
        if j < 2:
            continue
        if A.cell(row=j, column=coly).value == None:
            break
        # print(type(A.cell(row=j, column=coly).value))
        else:
            CH1_Current.append(abs(float(A.cell(row=j, column=coly).value)))
            CH2_Voltage.append(int(A.cell(row=j, column=colx).value))
        if j == i + 1:
            break

    return CH2_Voltage, CH1_Current

# 定义计算离散点导数的函数
def cal_deriv(x, y):  # x, y的类型均为列表
    diff_x = []  # 用来存储x列表中的两数之差
    for i, j in zip(x[0::], x[1::]):
        diff_x.append(j - i)

    diff_y = []  # 用来存储y列表中的两数之差
    for i, j in zip(y[0::], y[1::]):
        diff_y.append(j - i)

    slopes = []  # 用来存储斜率
    for i in range(len(diff_y)):
        slopes.append(diff_y[i] / diff_x[i])

    deriv = []  # 用来存储一阶导数
    for i, j in zip(slopes[0::], slopes[1::]):
        deriv.append(((i + j) / 2))  # 根据离散点导数的定义，计算并存储结果
    deriv.insert(0, slopes[0])  # (左)端点的导数即为与其最近点的斜率
    deriv.append(slopes[-1])  # (右)端点的导数即为与其最近点的斜率

    return deriv  # 返回存储一阶导数结果的列表

# 计算列表中全部数值的平均值
def cal_ave(list_num):

    i = 0
    sum = 0
    for each in list_num:
        sum += each
        i += 1
    ave = sum/i

    return ave

if __name__ == '__main__':

    # 存储每个比例的平均迁移率
    u = []
    I_onoff = []

    # 依次从 PS 质量分数为 0% 的数据读取到 PS 质量分数为 90% 的数据
    for i in range(0, 100, 10):
        files_name = load_files(i)
        print('PS:IDT-BT 共混比例为 %d : %d 的文件有：' %(i, 100-i))

        # 计算同一比例下文件数有多少个
        count_num = 0

        # 存储单个比例下的每个电极的迁移率，用于计算该比例的平均迁移率
        u_each = []
        I_onoff_each = []

        # 将同一比例下的数据名称输出
        for each_name in files_name:

            # 读取 excel 文件内部数据
            # param1 是电压数据，param2 是电流数据
            param1, param2 = load_excel(i, each_name)

            paramI = []
            for each_I in param2[10:41]:
                paramI.append(math.sqrt(each_I))
            # 电流开根号，对电压取微分，乘10000，平方后再除以5，选取第一遍扫的结果，选取其中最大值
            # for each in cal_deriv(param1[0:41], paramI):
            #    u.append((each * 10000)**2 / 5)

            # 计算迁移率
            deriv_u = []
            for each_result in cal_deriv(param1[10:41], paramI):
                deriv_u.append(abs(each_result))
            deriv_u.sort(reverse=True)
            u_each.append((deriv_u[0] * 10000) ** 2 / 5)

            # 取栅压对应为 20V 时作为关态电流，取栅压为 -60V 时作为开态电流，计算开关比
            param2_I = param2[0:41].copy()
            I_onoff_each.append(math.log(param2_I[40] / param2_I[9], 10))

            # 输出该文件夹下的数据文件名
            print(each_name)
            count_num += 1
        print('共 %d 个数据文件' %count_num)

        u.append(cal_ave(u_each))
        I_onoff.append(cal_ave(I_onoff_each))

        print('\n')

    p = 0
    print('迁移率为：')
    for each_u in u:
        print('PS:IDT-BT = ' + str(p) + ':'+ str(100-p) + '  ' + str(each_u))
        p += 10
    print('\n')

    p = 0
    print('开关比为：')
    for each_I_onoff in I_onoff:
        print('PS:IDT-BT = ' + str(p) + ':'+ str(100-p) + '  ' + str(each_I_onoff))
        p += 10

    print()

