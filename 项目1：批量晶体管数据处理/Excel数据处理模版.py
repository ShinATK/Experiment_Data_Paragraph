from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import os
import math
'''
自动生成晶体管输出曲线

栅压默认 20V 到 -60V
电流纵轴上下限，自动根据数据确定(目前只能按照设定好的几个数值作图)
将数据导出到 excel 表格后，放入指定的文件夹内 （例如：/Users/mac/Documents/Data/Excel/）
程序将自动作图并保存在 Data 文件夹下的 Excel_Fig 文件夹内，图表命名按照对应的 excel 表格。

'''

def load_files(n, path='S:/Xjtu/我的实验数据/PS IDT-BT/'):
    '''

    :param path:
    :return: files_xlsx
    '''
    # 读取路径默认，需要在调用时手动更改
    # 重复读取excel文件，并依次交给 load_excel 函数执行

    # 读取 path 下的所有文件的文件名
    files_name = os.listdir(path + str(n) + '/')
    # print(files_name)

    # 存储后缀名为xlsx的文件名称
    files_xlsx = []

    # 根据后缀名筛选excel表格文件
    for each in files_name:
        files_extension = os.path.splitext(each)[1]
        # print(files_extension)
        if files_extension == '.xlsx':
            files_xlsx.append(each)

    files_xlsx.sort()
    return files_xlsx


def load_excel(path, i=82, colx=7, coly=5):
    '''

    :param path:
    :param i:
    :param colx:
    :param coly:
    :return: CH1_Current, CH2_Voltage
    '''
    
    # 导入excel内所需要的数据
    # path 数据存放路径；i 导入几行(从第二行开始计数);
    # colx 导入的作为x轴的数据的列号;coly 导入的作为y轴的数据的列号
    # 部分数据存在少于81组数据的情况，需要修改 i 这部分的数值。不需要修改，
    # 这部分少数据大概率是因为软件导出到 excel 表格的时候出错，导致部分数据缺失了

    a = load_workbook(path)
    A = a.active

    current = []
    voltage = []

    for j in range(1, i+1):
        if j < 2:
            continue
        if A.cell(row=j, column=coly).value == None:
            break
        # print(type(A.cell(row=j, column=coly).value))
        else:
            current.append(abs(float(A.cell(row=j, column=coly).value)))
            voltage.append(int(A.cell(row=j, column=colx).value))
        if j == i+1:
            break

    return voltage, current


def draw_graph(x, y, n, excel_name):
    '''

    :param x:
    :param y:
    :param name;
    :return: 0
    '''
    # 利用导入的excel数据生成设定好的图像（一个）

    fig, ax = plt.subplots(1, 1)

    # 刻度调整为朝内
    # plt.rcParams['xtick.direction'] = 'out'  # 将x轴的刻度方向设置向内
    # plt.rcParams['ytick.direction'] = 'in'  # 将y轴的刻度方向设置向内

    plt.plot(x, y, linestyle = '--', color = 'green', marker='o')

    plt.xlim([-65, 25])

    # y轴的坐标值需要根据实际的实验数据自动确定
    ymin = [1e-4, 5e-5, 1e-5, 5e-6, 1e-6, 5e-7, 1e-7, 5e-8, 1e-8, 5e-9, 1e-9, 5e-10, 1e-10, 5e-11, 1e-11, 5e-12, 1e-12]
    ymax = [1e-4, 5e-5, 1e-5, 5e-6, 1e-6, 5e-7, 1e-7, 5e-8, 1e-8, 5e-9, 1e-9, 5e-10, 1e-10, 5e-11, 1e-11, 5e-12, 1e-12]

    ymin.sort(reverse=True) # ymin 需要升序排列，从而保证找出 ymin 中小于数据最小值中的最大值
    ymax.sort() # ymax 需要降序排列，从而保证找出 ymax 中的大于数据最大值中的最小值

    temp_min = min(y)
    temp_max = max(y)

    for each in ymin:
        if each <= temp_min:
            temp_min = each
            break
        else:
            continue

    for each in ymax:
        if each >= temp_max:
            temp_max = each
            break
        else:
            continue

    plt.ylim([temp_min, temp_max])

    my_x_ticks = np.arange(-60, 40, 20)
    plt.xticks(my_x_ticks)
    # plt.xticks([-60, -40, -20, 0, 20])

    # 纵轴修改为对数形式
    ax.semilogy()

    # 设定图像标题
    # plt.title(name)

    # 设定图像的轴坐标名称
    font1 = {'family' : 'Times New Roman',
            'weight' : 'normal',
            'style' : 'italic',
            'size'   : 14
            }



    plt.xlabel('V$_g$$_s$ (V)', font1)
    plt.ylabel('-I$_s$$_d$ (A)', font1)

    plt.title(excel_name[0:-5])

    # 这里的参数更改为图像存放位置
    plt.savefig('S:/Xjtu/我的实验数据/PS IDT-BT/'+ str(n) + '/' + excel_name.split('.', 1)[0] + '.png', dpi=720)
    # plt.draw()
    # plt.pause(0.2)
    plt.close()
    return 0


# 定义计算离散点导数的函数
def cal_deriv(x, y):  # x, y的类型均为列表
    diff_x = []  # 用来存储x列表中的两数之差
    for i, j in zip(x[0::], x[1::]):
        diff_x.append(j - i)

    diff_y = []  # 用来存储y列表中的两数之差
    for i, j in zip(y[0::], y[1::]):
        diff_y.append(j - i)

    slopes = []  # 用来存储斜率
    for i in range(len(diff_y)):
        slopes.append(diff_y[i] / diff_x[i])

    deriv = []  # 用来存储一阶导数
    for i, j in zip(slopes[0::], slopes[1::]):
        deriv.append(((i + j) / 2))  # 根据离散点导数的定义，计算并存储结果
    deriv.insert(0, slopes[0])  # (左)端点的导数即为与其最近点的斜率
    deriv.append(slopes[-1])  # (右)端点的导数即为与其最近点的斜率

    #for i in deriv:  # 打印结果，方便检查，调用时也可注释掉
    #     print(i)

    return deriv  # 返回存储一阶导数结果的列表


if __name__ == '__main__':

    for p in range(60, 100, 10):

        #p = 95
        print('PS:IDT-BT比例为 %d ：%d' % (p, 100-p))
        file_name = load_files(p)
        for each_name in file_name:
            param1, param2 = load_excel('S:/Xjtu/我的实验数据/PS IDT-BT/'+ str(p) + '/' + each_name)
            paramI = []

            for each in param2[10:41]:
                paramI.append(math.sqrt(each))

            deriv = []
            for each in cal_deriv(param1[10:41], paramI):
                deriv.append(abs(each))
            deriv.sort(reverse=True)

            # 求迁移率
            u = []
            for each in deriv[0:5]:
                u.append((each*10000)**2 / 5)

            # 求开关比
            param2_I = param2[0:41].copy()
            param2_I.sort(reverse=True)

            I_onoff = math.log(param2_I[0]/param2_I[-1], 10)


            print(each_name +'前5组迁移率为:%0.2f; %0.2f, %0.2f, %0.2f, %0.2f'
                                                    % (u[0], u[1], u[2], u[3], u[4]))
            print(each_name +'开关比为：%0.2f\n' % I_onoff)
            draw_graph(param1, param2, p, each_name)

    # input('\n\n按下 enter 键退出')
