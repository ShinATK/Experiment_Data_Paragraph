from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import os
'''

不同混合比例下, 每个片子的最大电流，开关比，迁移率

'''


def load_files(path='/Users/mac/Documents/Data'):
    '''
    获取excel表格
    :param path:
    :return: files_xlsx
    '''
    # 读取路径默认，需要手动更改
    # 重复读取excel文件，并依次交给 load_excel 函数执行

    # 读取 path 下的所有文件的文件名
    files_name = os.listdir(path)
    # print(files_name)

    # 存储后缀名为xlsx的文件名称
    files_xlsx = []

    # 根据后缀名筛选excel表格文件
    for each in files_name:
        files_extension = os.path.splitext(each)[1]
        # print(files_extension)
        if files_extension == '.xlsx':
            files_xlsx.append(each)

    return files_xlsx


def load_excel(path, i=82, colx=7, coly=5):
    '''

    :param path:
    :param i:
    :param colx:
    :param coly:
    :return: CH1_Current, CH2_Voltage
    '''
    # 导入excel内所需要的数据
    # path 数据存放路径；i 导入几行(从第二行开始计数);
    # colx 导入的作为x轴的数据的列号;coly 导入的作为y轴的数据的列号
    # 部分数据存在少于81组数据的情况，需要修改 i 这部分的数值。不需要修改，
    # 这部分少数据大概率是因为软件导出到 excel 表格的时候出错，导致部分数据缺失了

    a = load_workbook(path)
    A = a.active

    CH1_Current = []
    CH2_Voltage = []

    for j in range(1, i+1):
        if j < 2:
            continue
        if A.cell(row=j, column=coly).value == None:
            break
        # print(type(A.cell(row=j, column=coly).value))
        else:
            CH1_Current.append(abs(float(A.cell(row=j, column=coly).value)))
            CH2_Voltage.append(int(A.cell(row=j, column=colx).value))
        if j == i+1:
            break

    return CH2_Voltage, CH1_Current

def draw_graph(y,name):
    '''

    :param y:
    :param name;
    :return: 0
    '''
    # 利用导入的excel数据生成设定好的图像（一个）

    fig, ax = plt.subplots(1, 1)

    # 刻度调整为朝内
    # plt.rcParams['xtick.direction'] = 'out'  # 将x轴的刻度方向设置向内
    # plt.rcParams['ytick.direction'] = 'in'  # 将y轴的刻度方向设置向内

    # x = ['70:30', '60:40', '50:50', '40:60', '30:70', '20:80']
    x = range(60, 10)
    plt.plot(x, y, linestyle = '--', color = 'green', marker='o')
    my_x_ticks = np.arange(0, 50, 10)
    plt.xticks(my_x_ticks)
    # plt.xticks([-60, -40, -20, 0, 20])

    # 纵轴修改为对数形式
    ax.semilogy()

    # 设定图像标题
    # plt.title(name)

    # 设定图像的轴坐标名称
    font1 = {'family' : 'Times New Roman',
            'weight' : 'normal',
            'style' : 'italic',
            'size'   : 14
            }



    plt.xlabel('IDT-BT (%)', font1)
    plt.ylabel('-I$_s$$_d$ (A)', font1)

    plt.draw()
    plt.pause(0.5)


    # while True:
    #
    #     temp = input('是否人为设定上下限（Y/N）:')
    #     # 问题：如何重新设置上下限
    #     if temp == 'Y' or temp == 'y':
    #         y_min = input('请输入y轴最小值：')
    #         y_max = input('请输入y轴最大值：')
    #         plt.ylim([y_min, y_max])
    #         plt.close()
    #     elif temp == 'N' or temp == 'n':
    #         break
    #     else:
    #         print('输入错误')

    # plt.show()
    plt.savefig('/Users/mac/Documents/Data/Fig/' + name + '.png', dpi=720)
    plt.close()
    return 0


if __name__ == '__main__':
    file_name = load_files()
    each_max_current = []
    while True:
        i = 0
        for each_name in file_name:
            param1 = load_excel('/Users/mac/Documents/Data/' + each_name)
            each_max_current.append(param1)
            i += 1
            if i == 9:
                break
            # print(param1, '\n',param2)
            # draw_graph(param1, each_name)
    input('\n\n按下 enter 键退出')
